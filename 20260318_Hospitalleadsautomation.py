"""
Global Healthcare & Daycare News Scraper
Covers GCC, Anglophone Africa, Francophone Africa, and Caribbean markets.

Query architecture:
  Layer 1 — Universal queries   : run for every country (hospital/clinic/daycare baseline)
  Layer 2 — Cluster-specific    : additive market-aware queries on top of universal layer

Three Excel sheets:
  Sheet 1 — Healthcare News     : one row per unique news event
  Sheet 2 — Apollo Results      : one row per unique company (30-day cache)
  Sheet 3 — Processed Hashes   : every article ever fetched (30-day rolling)

Requirements:
    pip install feedparser openpyxl requests openai
"""

import difflib
import feedparser
import hashlib
import logging
import os
import re
import time
import requests
import openai
import json
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────

OPENAI_API_KEY   = "**************"
APOLLO_API_KEY   = "***************"
EXCEL_FILE_PATH  = os.path.expanduser("~/Desktop/GCC_Healthcare_News_0.18.xlsx")
LOG_FILE_PATH    = os.path.expanduser("~/Desktop/gcc_news.log")

ARTICLE_PAUSE        = 0
FETCH_TIMEOUT        = 5
APOLLO_MIN_GAP       = 0.15
APOLLO_MAX_RETRIES   = 2
SIMILARITY_THRESHOLD = 0.60
EVENT_DEDUP_DAYS     = 7
APOLLO_CACHE_DAYS    = 30
HASH_RETENTION_DAYS  = 30   # processed hashes older than this are purged

NAME_NOISE_SUFFIXES = [
    "academy", "foundation", "factory", "institute", "trust",
    "fund", "charity", "association", "society", "club",
    "llc", "ltd", "inc", "plc", "corp", "co", "wll", "fze", "llp",
]

FINGERPRINT_STRIP_WORDS = {
    "medical", "services", "service", "healthcare", "health", "care",
    "hospital", "hospitals", "clinic", "clinics", "group", "holding",
    "international", "national", "company", "corporation", "centre",
    "center", "the", "of", "and", "for", "al", "el",
}

# ─────────────────────────────────────────────
# COUNTRIES & MARKET CLUSTERS
# ─────────────────────────────────────────────

# Full country list
COUNTRIES = [
    # GCC
    "Saudi Arabia", "Kuwait", "Qatar", "Bahrain", "Oman",
    # Anglophone Africa
    "Ghana", "Tanzania", "Kenya", "Uganda", "Nigeria",
    "Botswana", "Zimbabwe", "Namibia", "Rwanda", "Mozambique", "Mauritius",
    # Francophone Africa
    "Burkina Faso", "Morocco",
    # Caribbean
    "Caribbean", "Trinidad and Tobago", "The Bahamas",
]

CATEGORIES = ["Hospital", "Clinic", "Daycare"]

# Market cluster assignment — determines which additive queries run
MARKET_CLUSTERS = {
    # GCC
    "Saudi Arabia":        "gcc",
    "Kuwait":              "gcc",
    "Qatar":               "gcc",
    "Bahrain":             "gcc",
    "Oman":                "gcc",
    # Anglophone Africa
    "Ghana":               "anglophone_africa",
    "Tanzania":            "anglophone_africa",
    "Kenya":               "anglophone_africa",
    "Uganda":              "anglophone_africa",
    "Nigeria":             "anglophone_africa",
    "Botswana":            "anglophone_africa",
    "Zimbabwe":            "anglophone_africa",
    "Namibia":             "anglophone_africa",
    "Rwanda":              "anglophone_africa",
    "Mozambique":          "anglophone_africa",
    "Mauritius":           "anglophone_africa",
    # Francophone Africa
    "Burkina Faso":        "francophone_africa",
    "Morocco":             "francophone_africa",
    # Caribbean
    "Caribbean":           "caribbean",
    "Trinidad and Tobago": "caribbean",
    "The Bahamas":         "caribbean",
}

# ─────────────────────────────────────────────
# LAYER 1 — UNIVERSAL QUERIES (all countries)
# Core facility types in standard English — guaranteed baseline for every market
# ─────────────────────────────────────────────

UNIVERSAL_QUERIES = {
    "Hospital": [
        'new hospital "{country}"',
        'hospital expansion "{country}"',
        'hospital construction "{country}"',
        'hospital investment "{country}"',
        'multispeciality hospital "{country}"',
        'specialist hospital "{country}"',
    ],
    "Clinic": [
        'new clinic "{country}"',
        'clinic opening "{country}"',
        'polyclinic "{country}"',
        'multispeciality clinic "{country}"',
        'medical centre opening "{country}"',
    ],
    "Daycare": [
        'new daycare "{country}"',
        'nursery opening "{country}"',
        'childcare centre "{country}"',
        'early childhood centre "{country}"',
    ],
}

# ─────────────────────────────────────────────
# LAYER 2 — CLUSTER-SPECIFIC ADDITIVE QUERIES
# Run on top of universal queries for each cluster
# ─────────────────────────────────────────────

CLUSTER_QUERIES = {

    "gcc": {
        "Hospital": [
            'hospital PPP "{country}"',
            'healthcare investment "{country}"',
        ],
        "Clinic": [
            'health city "{country}"',
            'outpatient centre "{country}"',
        ],
        "Daycare": [
            'nursery school "{country}"',
        ],
    },

    "anglophone_africa": {
        "Hospital": [
            'hospital PPP "{country}"',
            'hospital concession "{country}"',
            'World Bank hospital "{country}"',
            'AfDB health "{country}"',
            'IFC hospital "{country}"',
            'district hospital "{country}"',
            'specialist hospital "{country}"',
            'diagnostic centre "{country}"',
            'maternity centre "{country}"',
            'health facility "{country}"',
        ],
        "Clinic": [
            'health centre opening "{country}"',
            'diagnostic centre "{country}"',
            'outpatient clinic "{country}"',
            'medical facility "{country}"',
        ],
        "Daycare": [
            'creche opening "{country}"',
            'ECD centre "{country}"',
            'early childhood development "{country}"',
            'nursery school "{country}"',
        ],
    },

    "francophone_africa": {
        "Hospital": [
            # French language queries
            'nouvel hôpital "{country}"',
            'construction hôpital "{country}"',
            'investissement santé "{country}"',
            'hôpital PPP "{country}"',
            'centre hospitalier "{country}"',
            'clinique privée "{country}"',
            # English fallback for international coverage
            'hospital PPP "{country}"',
            'health facility "{country}"',
            'diagnostic centre "{country}"',
        ],
        "Clinic": [
            'nouvelle clinique "{country}"',
            'centre de santé "{country}"',
            'polyclinique "{country}"',
            'medical centre "{country}"',
        ],
        "Daycare": [
            'crèche "{country}"',
            'école maternelle "{country}"',
            'centre petite enfance "{country}"',
            'nursery school "{country}"',
        ],
    },

    "caribbean": {
        "Hospital": [
            'new hospital "{country}"',
            'hospital expansion "{country}"',
            'health facility "{country}"',
            'medical centre "{country}"',
            'hospital groundbreaking "{country}"',
        ],
        "Clinic": [
            'health centre "{country}"',
            'wellness centre "{country}"',
            'medical centre opening "{country}"',
            'outpatient centre "{country}"',
        ],
        "Daycare": [
            'nursery school "{country}"',
            'early childhood centre "{country}"',
            'infant school "{country}"',
            'creche "{country}"',
        ],
    },
}

VAGUE_NAME_TRIGGERS = [
    "ministry", "government", "authority", "department", "municipality",
    "council", "committee", "board", "unknown", "developer", "contractor",
    "company", "group", "firm", "local", "private", "public", "n/a", "none",
]

# ─────────────────────────────────────────────
# TITLE SCORING
# ─────────────────────────────────────────────

APOLLO_FETCH_TITLES = [
    "CEO", "Chief Executive Officer", "Chief Executive",
    "Managing Director", "MD", "Chairman", "Vice Chairman",
    "President", "Vice President", "COO", "Chief Operating Officer",
    "CMO", "Chief Medical Officer", "CFO", "Chief Financial Officer",
    "CSO", "Chief Strategy Officer", "Chief Development Officer",
    "Director", "Executive Director", "Group Director",
    "Director of Operations", "Director of Development",
    "Director of Facilities", "Director of Projects",
    "Director of Healthcare", "Director of Procurement",
    "Director of Business Development", "Project Director",
    "Commercial Director", "Investment Director", "Portfolio Director",
    "VP", "Vice President of Operations",
    "Head of Operations", "Head of Development", "Head of Projects",
    "Head of Healthcare", "Head of Procurement", "Head of Facilities",
    "Head of Business Development", "General Manager", "GM",
    "Regional Manager", "Country Manager", "Centre Manager",
    "Branch Manager", "Owner", "Co-Owner", "Proprietor",
    "Founder", "Co-Founder", "Operations Manager", "Project Manager",
    "Partner", "Managing Partner", "Contracts Manager",
]

SENIORITY_SCORES = {
    5: ["ceo","chief executive","chief executive officer","chairman","vice chairman",
        "managing director","md","president","founder","co-founder","owner","partner",
        "managing partner"],
    4: ["coo","chief operating officer","cmo","chief medical officer","cfo",
        "chief financial officer","cso","chief strategy officer",
        "chief development officer","executive director","group director",
        "investment director","portfolio director","commercial director",
        "vice president","vp"],
    3: ["director","director of operations","director of development",
        "director of facilities","director of projects","director of healthcare",
        "director of procurement","director of business development",
        "project director","regional manager","country manager",
        "general manager","gm","head of operations","head of development",
        "head of projects","head of healthcare","head of procurement",
        "head of facilities","head of business development"],
    2: ["manager","centre manager","branch manager","operations manager",
        "project manager","contracts manager","co-owner","proprietor"],
    1: [],
}

RELEVANCE_BONUS = {
    2: ["development","projects","project","facilities","facility","procurement",
        "construction","infrastructure","operations","healthcare","health care",
        "medical","clinical","business development","strategy","investment","portfolio"],
    1: ["commercial","contracts","regional","country","general manager",
        "managing director","chief executive","chairman","founder","owner"],
    0: [],
}

def score_contact(title):
    t = title.lower().strip()
    seniority = 1
    for score, kws in SENIORITY_SCORES.items():
        if any(kw in t for kw in kws):
            seniority = score
            break
    bonus = 0
    for b, kws in RELEVANCE_BONUS.items():
        if any(kw in t for kw in kws):
            bonus = b
            break
    return seniority + bonus

def prioritise_contacts(people_raw, logger):
    scored = [(score_contact(p.get("title","")), p) for p in people_raw]
    scored.sort(key=lambda x: x[0], reverse=True)
    if scored:
        logger.info(f"Top contact: '{scored[0][1].get('name')}' "
                    f"| '{scored[0][1].get('title')}' | score {scored[0][0]}")
    return [p for _, p in scored[:3]]

# ─────────────────────────────────────────────
# HEADERS
# ─────────────────────────────────────────────

NEWS_HEADERS = [
    "Date Found", "Headline", "Source", "URL", "Country",
    "Category", "News Type", "AI Summary", "Relevance Score",
    "Relevance Reason", "Key Signal", "Management Name",
    "Company Website", "Company LinkedIn", "Company Size",
    "Company Description",
    "Key Contact 1 — Name", "Key Contact 1 — Title", "Key Contact 1 — Score",
    "Key Contact 1 — Email", "Key Contact 1 — Phone", "Key Contact 1 — LinkedIn",
    "Key Contact 2 — Name", "Key Contact 2 — Title", "Key Contact 2 — Score",
    "Key Contact 2 — Email", "Key Contact 2 — Phone", "Key Contact 2 — LinkedIn",
    "Key Contact 3 — Name", "Key Contact 3 — Title", "Key Contact 3 — Score",
    "Key Contact 3 — Email", "Key Contact 3 — Phone", "Key Contact 3 — LinkedIn",
    "Apollo Match Confidence", "Research Flag", "Event Fingerprint", "Hash",
]

APOLLO_HEADERS = [
    "Management Name", "Country", "Last Looked Up", "Apollo Match Confidence",
    "Company Website", "Company LinkedIn", "Company Size", "Company Description",
    "Key Contact 1 — Name", "Key Contact 1 — Title", "Key Contact 1 — Score",
    "Key Contact 1 — Email", "Key Contact 1 — Phone", "Key Contact 1 — LinkedIn",
    "Key Contact 2 — Name", "Key Contact 2 — Title", "Key Contact 2 — Score",
    "Key Contact 2 — Email", "Key Contact 2 — Phone", "Key Contact 2 — LinkedIn",
    "Key Contact 3 — Name", "Key Contact 3 — Title", "Key Contact 3 — Score",
    "Key Contact 3 — Email", "Key Contact 3 — Phone", "Key Contact 3 — LinkedIn",
]

HASH_HEADERS = ["Hash", "Date", "Outcome"]   # Sheet 3

NEWS_TYPE_GUIDE = """
Classify into exactly one:
- "New Construction"      : Greenfield — brand new facility being planned or built
- "Expansion"             : Existing facility adding capacity, new wing, beds, or floor
- "Acquisition / JV"      : Ownership change, merger, management takeover, or new JV
- "Regulatory / Licensing": Government approval, permit, accreditation — early-stage signal
- "General News"          : Operational updates, staff changes, awards — low outreach priority
"""

RELEVANCE_GUIDE = """
Score using this checklist. Count how many signals are present:

UNIVERSAL SIGNALS (apply to all markets):
  [1] Named decision-maker, executive, or project owner identified
  [2] Specific investment figure or contract value mentioned
  [3] Confirmed opening date, completion year, or timeline stated
  [4] Active construction or development confirmed (not just planned/rumoured)
  [5] Bed capacity, floor area, or facility size mentioned

AFRICA-SPECIFIC SIGNALS (count as signals for African country articles):
  [6] PPP award or concession agreement signed for a named facility
  [7] World Bank / AfDB / IFC / donor funding confirmed for a named facility
  [8] Government contract awarded to a named private operator or developer

CARIBBEAN-SPECIFIC SIGNALS (count as signals for Caribbean country articles):
  [9] Groundbreaking, sod turning, or ribbon cutting for a named facility
  [10] Government or regional health authority announcement of a named new centre

Rules:
- "High"   : 2 or more signals present
- "Medium" : exactly 1 signal, OR named facility in a broader development story
- "Low"    : 0 signals — passing mention, general commentary, not actionable
"""

NEWS_TYPE_COLORS = {
    "New Construction":       ("D9EAD3", "274E13"),
    "Expansion":              ("CFE2F3", "1C4587"),
    "Acquisition / JV":       ("FFF2CC", "7F6000"),
    "Regulatory / Licensing": ("FCE5CD", "783F04"),
    "General News":           ("F3F3F3", "444444"),
}

CATEGORY_COLORS = {
    "Hospital":               ("D9EAD3", "274E13"),
    "Clinic":                 ("CFE2F3", "1C4587"),
    "Multispeciality Clinic": ("EAD1DC", "4A1942"),
    "Daycare":                ("FFF2CC", "7F6000"),
    "Other":                  ("F3F3F3", "444444"),
}
RELEVANCE_COLORS = {
    "High":   ("C6EFCE", "276221"),
    "Medium": ("FFEB9C", "9C5700"),
    "Low":    ("F4CCCC", "990000"),
}
FLAG_COLORS = {
    "OK":               None,
    "PARTIAL DATA":     ("FFF2CC", "7F6000"),
    "MANUAL RESEARCH":  ("FCE4D6", "843C0C"),
    "NOT IN APOLLO":    ("EAD1DC", "4A1942"),
    "APOLLO MISMATCH":  ("F4CCCC", "990000"),
}

# ─────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────

def setup_logging():
    logger = logging.getLogger("gcc_scraper")
    logger.setLevel(logging.DEBUG)
    fh = logging.FileHandler(LOG_FILE_PATH, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s",
                            "%Y-%m-%d %H:%M:%S")
    fh.setFormatter(fmt)
    ch.setFormatter(fmt)
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger

# ─────────────────────────────────────────────
# NAME / FINGERPRINT UTILITIES
# ─────────────────────────────────────────────

def normalise_for_fingerprint(name):
    if not name:
        return ""
    clean = re.sub(r"[^\w\s]", "", name.lower().strip())
    return " ".join(w for w in clean.split() if w not in FINGERPRINT_STRIP_WORDS)

def make_event_fingerprint(management_name, country, news_type):
    norm = normalise_for_fingerprint(management_name)
    return f"{norm}|{country.lower()}|{news_type.lower()}"

def apollo_cache_key(management_name, country):
    return f"{normalise_for_fingerprint(management_name)}|{country.lower()}"

def clean_management_name(name):
    if not name:
        return name, False
    words = name.strip().split()
    cleaned, stripped = [], False
    for word in words:
        if word.lower() in NAME_NOISE_SUFFIXES:
            stripped = True
            break
        cleaned.append(word)
    result = " ".join(cleaned).strip()
    return (result, stripped) if len(result.split()) >= 2 else (name, False)

def extract_domain(url):
    if not url:
        return ""
    try:
        d = re.sub(r"^https?://", "", url.lower())
        d = re.sub(r"^www\.", "", d)
        return d.split("/")[0].split("?")[0]
    except Exception:
        return ""

def is_vague_name(name):
    if not name or len(name.strip()) < 4:
        return True
    words = [w.strip().lower() for w in name.split() if w.strip()]
    vague = sum(1 for w in words if w in VAGUE_NAME_TRIGGERS)
    return vague >= len(words)

def name_similarity(a, b):
    a_c = re.sub(r"[^\w\s]", "", a.lower().strip())
    b_c = re.sub(r"[^\w\s]", "", b.lower().strip())
    if a_c in b_c or b_c in a_c:
        return 0.85
    noise = {"the","of","and","for","a","an","co","company","group","ltd","llc",
             "inc","plc","corp","medical","health","hospital","care","clinic",
             "centre","center","services","international","al","el"}
    a_sig = set(a_c.split()) - noise
    b_sig = set(b_c.split()) - noise
    if a_sig and b_sig:
        overlap = len(a_sig & b_sig) / max(len(a_sig), len(b_sig))
        if overlap >= 0.5:
            return max(overlap, 0.70)
    return difflib.SequenceMatcher(None, a_c, b_c).ratio()

def sanitise(val):
    """Convert None / literal 'None' / 'null' / 'N/A' to empty string."""
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s in ("None", "null", "N/A", "nan") else s

def contact_has_data(person):
    """Gap 5 fix — contact is considered found if ANY key field is populated."""
    return any(sanitise(person.get(f)) for f in ("name", "title", "email"))

# ─────────────────────────────────────────────
# APOLLO RATE-AWARE CALLER
# ─────────────────────────────────────────────

_last_apollo_call = 0.0

def _apollo_request(method, endpoint, logger, payload=None, params=None):
    global _last_apollo_call
    url = f"https://api.apollo.io/api/v1/{endpoint}"
    headers = {
        "Content-Type": "application/json",
        "Cache-Control": "no-cache",
        "X-Api-Key": APOLLO_API_KEY,
    }
    for attempt in range(1, APOLLO_MAX_RETRIES + 1):
        elapsed = time.time() - _last_apollo_call
        if elapsed < APOLLO_MIN_GAP:
            time.sleep(APOLLO_MIN_GAP - elapsed)
        try:
            if method == "GET":
                resp = requests.get(url, headers=headers, params=params,
                                    timeout=FETCH_TIMEOUT)
            else:
                resp = requests.post(url, headers=headers,
                                     json=payload, params=params,
                                     timeout=FETCH_TIMEOUT)
            _last_apollo_call = time.time()
            remaining = resp.headers.get("X-RateLimit-Remaining")
            if remaining is not None:
                logger.debug(f"Apollo remaining: {remaining}")
                if int(remaining) < 10:
                    logger.warning(f"Apollo rate limit low ({remaining}) — pausing 15s")
                    time.sleep(15)
            if resp.status_code == 429:
                retry_after = int(resp.headers.get("Retry-After", 60))
                logger.warning(f"Apollo 429 — waiting {retry_after}s (attempt {attempt})")
                time.sleep(retry_after)
                continue
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.HTTPError as e:
            logger.error(f"Apollo request error: {e}")
            if attempt < APOLLO_MAX_RETRIES:
                time.sleep(5 * attempt)
            else:
                return {}
        except requests.exceptions.RequestException as e:
            logger.error(f"Apollo connection error: {e}")
            time.sleep(5 * attempt)
    return {}

def apollo_post(endpoint, payload, logger):
    return _apollo_request("POST", endpoint, logger, payload=payload)

def apollo_get(endpoint, params, logger):
    return _apollo_request("GET", endpoint, logger, params=params)

# ─────────────────────────────────────────────
# GAP 2 — ORGANISATION ENRICHMENT
# ─────────────────────────────────────────────

def apollo_enrich_organisation(domain, logger):
    """
    Fetch full company profile via Organisation Enrichment endpoint.
    Returns size and description not available from search endpoint.
    """
    if not domain:
        return {}
    logger.info(f"Apollo org enrichment: {domain}")
    data = apollo_get("organizations/enrich", {"domain": domain}, logger)
    org  = data.get("organization", {})
    if not org:
        return {}
    return {
        "size":        sanitise(org.get("estimated_num_employees") or
                                org.get("employee_count")),
        "description": sanitise((org.get("short_description") or
                                 org.get("seo_description") or ""))[:300],
    }

# ─────────────────────────────────────────────
# GAP 1, 3, 4 — PEOPLE ENRICHMENT
# ─────────────────────────────────────────────

def apollo_enrich_person(person_id, logger):
    """
    Full people enrichment via bulk_match — correct endpoint for ID-based lookup.
    reveal_phone_number passed as query param per Apollo docs, not in body.
    Returns name, email, phone, LinkedIn.
    """
    if not person_id or str(person_id).strip() in ("", "None", "null"):
        logger.debug("apollo_enrich_person: skipped — no valid person_id")
        return {}
    logger.debug(f"Apollo person enrichment: {person_id}")

    global _last_apollo_call
    url = "https://api.apollo.io/api/v1/people/bulk_match"
    headers = {
        "Content-Type": "application/json",
        "Cache-Control": "no-cache",
        "X-Api-Key": APOLLO_API_KEY,
    }
    # reveal_phone_number must be a query param, not in body
    params  = {"reveal_personal_emails": "false", "reveal_phone_number": "false"}
    payload = {"details": [{"id": person_id}]}

    for attempt in range(1, APOLLO_MAX_RETRIES + 1):
        elapsed = time.time() - _last_apollo_call
        if elapsed < APOLLO_MIN_GAP:
            time.sleep(APOLLO_MIN_GAP - elapsed)
        try:
            resp = requests.post(url, headers=headers, params=params,
                                 json=payload, timeout=FETCH_TIMEOUT)
            _last_apollo_call = time.time()
            if resp.status_code == 429:
                retry_after = int(resp.headers.get("Retry-After", 60))
                logger.warning(f"Apollo enrich 429 — waiting {retry_after}s")
                time.sleep(retry_after)
                continue
            if resp.status_code == 400:
                logger.error(
                    f"Apollo bulk_match 400 — body: {resp.text[:300]}"
                )
                return {}
            resp.raise_for_status()
            matches = resp.json().get("matches", [])
            if not matches:
                return {}
            person = matches[0]

            name     = sanitise(person.get("name"))
            email    = sanitise(person.get("email"))
            linkedin = sanitise(person.get("linkedin_url"))
            phone    = ""
            phone_numbers = person.get("phone_numbers", [])
            if phone_numbers:
                phone = sanitise(phone_numbers[0].get("sanitized_number") or
                                 phone_numbers[0].get("raw_number"))
            if not phone:
                phone = sanitise(person.get("sanitized_phone"))

            return {"name": name, "email": email,
                    "linkedin": linkedin, "phone": phone}

        except requests.exceptions.HTTPError as e:
            logger.error(f"Apollo enrich error: {e}")
            if attempt < APOLLO_MAX_RETRIES:
                time.sleep(5 * attempt)
            else:
                return {}
        except Exception as e:
            logger.error(f"Apollo enrich connection error: {e}")
            time.sleep(5 * attempt)
    return {}

# ─────────────────────────────────────────────
# APOLLO COMPANY SEARCH
# ─────────────────────────────────────────────

def prepare_apollo_input(management_name, website, country, logger):
    cleaned_name, was_stripped = clean_management_name(management_name)
    if was_stripped:
        logger.info(f"Name cleaned: '{management_name}' -> '{cleaned_name}'")
    domain = extract_domain(website)
    payloads = []
    if domain:
        payloads.append({
            "strategy": "domain",
            "payload": {"q_organization_domains": domain,
                        "page": 1, "per_page": 1}
        })
    payloads.append({
        "strategy": "name+keywords+country",
        "payload": {
            "q_organization_name": cleaned_name,
            "organization_locations": [country],
            "q_keywords": "hospital clinic healthcare medical",
            "page": 1, "per_page": 5,
        }
    })
    payloads.append({
        "strategy": "name+keywords",
        "payload": {
            "q_organization_name": cleaned_name,
            "q_keywords": "hospital clinic healthcare medical",
            "page": 1, "per_page": 5,
        }
    })
    payloads.append({
        "strategy": "name only",
        "payload": {"q_organization_name": cleaned_name,
                    "page": 1, "per_page": 5}
    })
    return payloads, cleaned_name


def apollo_company_search(management_name, website, country, logger):
    payloads, cleaned_name = prepare_apollo_input(
        management_name, website, country, logger
    )
    for item in payloads:
        strategy = item["strategy"]
        data     = apollo_post("mixed_companies/search", item["payload"], logger)
        orgs     = data.get("organizations", [])
        if not orgs:
            logger.info(f"Apollo [{strategy}]: no results")
            continue
        if strategy == "domain":
            org = orgs[0]
            logger.info(f"Apollo [domain] ACCEPTED: '{org.get('name')}'")
            return org, 1.0, strategy
        best_org, best_score = None, 0.0
        for org in orgs:
            score = name_similarity(cleaned_name, org.get("name", ""))
            logger.info(
                f"Apollo [{strategy}] candidate: "
                f"'{org.get('name')}' | score: {score:.2f}"
            )
            if score > best_score:
                best_score, best_org = score, org
        if best_score >= SIMILARITY_THRESHOLD:
            logger.info(
                f"Apollo [{strategy}] ACCEPTED: "
                f"'{best_org.get('name')}' ({best_score:.2f})"
            )
            return best_org, round(best_score, 2), strategy
        logger.warning(
            f"Apollo [{strategy}] REJECTED: "
            f"best {best_score:.2f} < {SIMILARITY_THRESHOLD}"
        )
    logger.info(f"Apollo: all strategies exhausted for '{cleaned_name}'")
    return None, 0.0, None


def apollo_people_search(company_id, company_name, country, logger):
    global _last_apollo_call
    logger.info(f"Apollo people search: {company_name} | {country}")
    url = "https://api.apollo.io/api/v1/mixed_people/api_search"
    headers = {
        "Content-Type": "application/json",
        "Cache-Control": "no-cache",
        "X-Api-Key": APOLLO_API_KEY,
    }
    params = [("organization_ids[]", company_id)]
    for title in APOLLO_FETCH_TITLES:
        params.append(("person_titles[]", title))
    params += [("person_locations[]", country), ("page", 1), ("per_page", 10)]

    for attempt in range(1, APOLLO_MAX_RETRIES + 1):
        elapsed = time.time() - _last_apollo_call
        if elapsed < APOLLO_MIN_GAP:
            time.sleep(APOLLO_MIN_GAP - elapsed)
        try:
            resp = requests.post(url, headers=headers,
                                 params=params, timeout=FETCH_TIMEOUT)
            _last_apollo_call = time.time()
            if resp.status_code == 429:
                retry_after = int(resp.headers.get("Retry-After", 60))
                time.sleep(retry_after)
                continue
            resp.raise_for_status()
            people = resp.json().get("people", [])
            if not people:
                logger.info("No people with country filter — retrying without")
                params_no_loc = [(k,v) for k,v in params
                                 if k != "person_locations[]"]
                resp2 = requests.post(url, headers=headers,
                                      params=params_no_loc,
                                      timeout=FETCH_TIMEOUT)
                people = resp2.json().get("people", [])
            logger.info(f"Apollo people: {len(people)} raw contacts")
            return people
        except requests.exceptions.HTTPError as e:
            logger.error(f"Apollo people error: {e}")
            if attempt < APOLLO_MAX_RETRIES:
                time.sleep(5 * attempt)
            else:
                return []
        except Exception as e:
            logger.error(f"Apollo people connection error: {e}")
            time.sleep(5 * attempt)
    return []


def extract_company_fields(org):
    if not org:
        return {}
    return {
        "name":     sanitise(org.get("name")),
        "website":  sanitise(org.get("website_url") or org.get("primary_domain")),
        "linkedin": sanitise(org.get("linkedin_url")),
        "size":     "",         # populated separately via org enrichment
        "description": "",      # populated separately via org enrichment
    }


def extract_person_fields(person, logger):
    """
    Gaps 1, 3, 4 — call full People Enrichment for name, LinkedIn, phone.
    Falls back to search-tier data if enrichment fails.
    """
    person_id = person.get("id")
    title     = sanitise(person.get("title"))
    score     = score_contact(title)

    # Start with what the search endpoint returned
    name     = sanitise(person.get("name"))
    email    = sanitise(person.get("email"))
    linkedin = sanitise(person.get("linkedin_url"))
    phone    = ""
    phones   = person.get("phone_numbers", [])
    if phones:
        phone = sanitise(phones[0].get("sanitized_number") or
                         phones[0].get("raw_number"))

    # Log exactly what search tier returned
    logger.debug(
        f"Search-tier person — id: {person_id} | name: '{name}' | "
        f"title: '{title}' | email: '{email}'"
    )

    # Full enrichment — only if we have a valid Apollo person ID
    valid_id = (person_id and
                str(person_id).strip() not in ("", "None", "null"))
    if valid_id:
        enriched = apollo_enrich_person(person_id, logger)
        if enriched.get("name"):
            name = enriched["name"]
        if enriched.get("email"):
            email = enriched["email"]
        if enriched.get("linkedin"):
            linkedin = enriched["linkedin"]
        if enriched.get("phone"):
            phone = enriched["phone"]
    else:
        logger.debug(f"Skipping enrichment — invalid person_id: '{person_id}'")

    return {
        "name": name, "title": title, "score": score,
        "email": email, "phone": phone, "linkedin": linkedin,
    }

# ─────────────────────────────────────────────
# EXCEL — THREE SHEETS
# ─────────────────────────────────────────────

def get_or_create_workbook():
    if os.path.exists(EXCEL_FILE_PATH):
        wb = load_workbook(EXCEL_FILE_PATH)
        ws_news   = wb["GCC News"] if "GCC News" in wb.sheetnames else wb.active
        ws_apollo = (wb["Apollo Results"]
                     if "Apollo Results" in wb.sheetnames else None)
        ws_hashes = (wb["Processed Hashes"]
                     if "Processed Hashes" in wb.sheetnames else None)
        if ws_apollo is None:
            ws_apollo = wb.create_sheet("Apollo Results")
            _init_apollo_sheet(ws_apollo)
        if ws_hashes is None:
            ws_hashes = wb.create_sheet("Processed Hashes")
            _init_hash_sheet(ws_hashes)
        wb.save(EXCEL_FILE_PATH)
    else:
        wb        = Workbook()
        ws_news   = wb.active
        ws_news.title = "GCC News"
        _init_news_sheet(ws_news)
        ws_apollo = wb.create_sheet("Apollo Results")
        _init_apollo_sheet(ws_apollo)
        ws_hashes = wb.create_sheet("Processed Hashes")
        _init_hash_sheet(ws_hashes)
        wb.save(EXCEL_FILE_PATH)
    return wb, ws_news, ws_apollo, ws_hashes


def _init_news_sheet(ws):
    ws.append(NEWS_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
    col_widths = [
        12,55,20,40,15,12,22,60,14,45,45,
        28,28,36,16,40,
        28,28,8,30,18,36,
        28,28,8,30,18,36,
        28,28,8,30,18,36,
        20,28,30,35,
    ]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1,i).column_letter].width = w


def _init_apollo_sheet(ws):
    ws.append(APOLLO_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A3C5E")
    col_widths = [
        28,15,14,20,28,36,12,40,
        28,28,8,30,18,36,
        28,28,8,30,18,36,
        28,28,8,30,18,36,
    ]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1,i).column_letter].width = w


def _init_hash_sheet(ws):
    ws.append(HASH_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="404040")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18

# ─────────────────────────────────────────────
# GAP 10 — PROCESSED HASHES SHEET
# ─────────────────────────────────────────────

def load_processed_hashes(ws_news, ws_hashes, logger):
    """
    Load all previously seen hashes from:
      - Sheet 1 (written articles)
      - Sheet 3 (all processed articles including Low/discarded)
    Purge Sheet 3 entries older than HASH_RETENTION_DAYS.
    Returns combined set of hash strings.
    """
    hashes  = set()
    cutoff  = datetime.now() - timedelta(days=HASH_RETENTION_DAYS)
    today   = datetime.now().strftime("%Y-%m-%d")

    # Sheet 1 hashes
    hash_col = NEWS_HEADERS.index("Hash") + 1
    for row in ws_news.iter_rows(min_row=2, min_col=hash_col,
                                  max_col=hash_col, values_only=True):
        if row[0]:
            hashes.add(row[0])

    # Sheet 3 hashes — with age-based purge
    rows_to_keep = [list(ws_hashes[1])]   # keep header
    purged = 0
    for row in ws_hashes.iter_rows(min_row=2, values_only=True):
        h, date_val, outcome = row[0], row[1], row[2]
        if not h:
            continue
        try:
            row_date = (datetime.strptime(date_val, "%Y-%m-%d")
                        if isinstance(date_val, str) else (date_val or datetime.min))
        except Exception:
            row_date = datetime.min
        if row_date >= cutoff:
            hashes.add(h)
            rows_to_keep.append([h, date_val, outcome])
        else:
            purged += 1

    # Rewrite Sheet 3 without expired rows
    if purged > 0:
        ws_hashes.delete_rows(1, ws_hashes.max_row)
        ws_hashes.append(HASH_HEADERS)
        for cell in ws_hashes[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="404040")
        for row_data in rows_to_keep[1:]:
            ws_hashes.append(row_data)
        logger.info(f"Processed Hashes: purged {purged} entries older than {HASH_RETENTION_DAYS} days")

    logger.info(f"Total known hashes loaded: {len(hashes)} "
                f"(Sheet1 + Sheet3 combined)")
    return hashes


def write_processed_hash(wb, ws_hashes, item_hash, outcome, logger):
    """Write a hash to Sheet 3 immediately after processing."""
    try:
        ws_hashes.append([item_hash, datetime.now().strftime("%Y-%m-%d"), outcome])
        wb.save(EXCEL_FILE_PATH)
    except Exception as e:
        logger.error(f"Failed to write processed hash: {e}")

# ─────────────────────────────────────────────
# APOLLO RESULTS SHEET CACHE
# ─────────────────────────────────────────────

def load_apollo_cache(ws_apollo, logger):
    cache, row_index = {}, {}
    if ws_apollo is None:
        return cache, row_index
    cutoff = datetime.now() - timedelta(days=APOLLO_CACHE_DAYS)
    try:
        for i, row in enumerate(
            ws_apollo.iter_rows(min_row=2, values_only=True), start=2
        ):
            mgmt_name   = row[APOLLO_HEADERS.index("Management Name")]
            country     = row[APOLLO_HEADERS.index("Country")]
            last_lookup = row[APOLLO_HEADERS.index("Last Looked Up")]
            if not mgmt_name or not country:
                continue
            key = apollo_cache_key(mgmt_name, country)
            row_index[key] = i
            try:
                lookup_date = (
                    datetime.strptime(last_lookup, "%Y-%m-%d")
                    if isinstance(last_lookup, str)
                    else (last_lookup or datetime.min)
                )
            except Exception:
                lookup_date = datetime.min
            if lookup_date >= cutoff:
                cache[key] = {
                    "company": {
                        "name":        sanitise(mgmt_name),
                        "website":     sanitise(row[APOLLO_HEADERS.index("Company Website")]),
                        "linkedin":    sanitise(row[APOLLO_HEADERS.index("Company LinkedIn")]),
                        "size":        sanitise(row[APOLLO_HEADERS.index("Company Size")]),
                        "description": sanitise(row[APOLLO_HEADERS.index("Company Description")]),
                    },
                    "confidence": sanitise(row[APOLLO_HEADERS.index("Apollo Match Confidence")]),
                    "people": _extract_people_from_apollo_row(row),
                }
        logger.info(f"Apollo cache: {len(cache)} fresh / "
                    f"{len(row_index)-len(cache)} stale")
    except Exception as e:
        logger.error(f"Error loading Apollo cache: {e}")
    return cache, row_index


def _extract_people_from_apollo_row(row):
    slots = [
        ("Key Contact 1 — Name","Key Contact 1 — Title","Key Contact 1 — Score",
         "Key Contact 1 — Email","Key Contact 1 — Phone","Key Contact 1 — LinkedIn"),
        ("Key Contact 2 — Name","Key Contact 2 — Title","Key Contact 2 — Score",
         "Key Contact 2 — Email","Key Contact 2 — Phone","Key Contact 2 — LinkedIn"),
        ("Key Contact 3 — Name","Key Contact 3 — Title","Key Contact 3 — Score",
         "Key Contact 3 — Email","Key Contact 3 — Phone","Key Contact 3 — LinkedIn"),
    ]
    people = []
    for n,t,s,e,p,l in slots:
        people.append({
            "name":     sanitise(row[APOLLO_HEADERS.index(n)]),
            "title":    sanitise(row[APOLLO_HEADERS.index(t)]),
            "score":    sanitise(row[APOLLO_HEADERS.index(s)]),
            "email":    sanitise(row[APOLLO_HEADERS.index(e)]),
            "phone":    sanitise(row[APOLLO_HEADERS.index(p)]),
            "linkedin": sanitise(row[APOLLO_HEADERS.index(l)]),
        })
    return people


def write_apollo_cache_row(wb, ws_apollo, management_name, country,
                           company_fields, people, confidence,
                           existing_row_index, logger):
    today = datetime.now().strftime("%Y-%m-%d")
    while len(people) < 3:
        people.append({"name":"","title":"","score":"",
                       "email":"","phone":"","linkedin":""})
    row_data = [
        management_name, country, today, confidence,
        company_fields.get("website",""),
        company_fields.get("linkedin",""),
        company_fields.get("size",""),
        company_fields.get("description",""),
        people[0]["name"], people[0]["title"], people[0].get("score",""),
        people[0]["email"], people[0]["phone"], people[0]["linkedin"],
        people[1]["name"], people[1]["title"], people[1].get("score",""),
        people[1]["email"], people[1]["phone"], people[1]["linkedin"],
        people[2]["name"], people[2]["title"], people[2].get("score",""),
        people[2]["email"], people[2]["phone"], people[2]["linkedin"],
    ]
    key = apollo_cache_key(management_name, country)
    if key in existing_row_index:
        row_num = existing_row_index[key]
        for col_idx, val in enumerate(row_data, start=1):
            ws_apollo.cell(row=row_num, column=col_idx, value=val)
        logger.info(f"Apollo Results: updated '{management_name}'")
    else:
        ws_apollo.append(row_data)
        existing_row_index[key] = ws_apollo.max_row
        logger.info(f"Apollo Results: new entry '{management_name}'")
    try:
        wb.save(EXCEL_FILE_PATH)
    except Exception as e:
        logger.error(f"Failed to save Apollo Results: {e}")

# ─────────────────────────────────────────────
# NEWS SHEET WRITE
# ─────────────────────────────────────────────

def write_news_row(wb, ws, row, logger):
    try:
        ws.append(row)
        idx      = ws.max_row
        rel_col  = NEWS_HEADERS.index("Relevance Score") + 1
        nt_col   = NEWS_HEADERS.index("News Type") + 1
        flag_col = NEWS_HEADERS.index("Research Flag") + 1

        rel_cell  = ws.cell(row=idx, column=rel_col)
        nt_cell   = ws.cell(row=idx, column=nt_col)
        flag_cell = ws.cell(row=idx, column=flag_col)

        if rel_cell.value in RELEVANCE_COLORS:
            bg, fg = RELEVANCE_COLORS[rel_cell.value]
            rel_cell.fill = PatternFill("solid", fgColor=bg)
            rel_cell.font = Font(bold=True, color=fg)

        # Category colour coding
        cat_col  = NEWS_HEADERS.index("Category") + 1
        cat_cell = ws.cell(row=idx, column=cat_col)
        if cat_cell.value in CATEGORY_COLORS:
            bg, fg = CATEGORY_COLORS[cat_cell.value]
            cat_cell.fill = PatternFill("solid", fgColor=bg)
            cat_cell.font = Font(bold=True, color=fg)

        if nt_cell.value in NEWS_TYPE_COLORS:
            bg, fg = NEWS_TYPE_COLORS[nt_cell.value]
            nt_cell.fill = PatternFill("solid", fgColor=bg)
            nt_cell.font = Font(bold=True, color=fg)

        flag_val = flag_cell.value or "OK"
        for key, colors in FLAG_COLORS.items():
            if key in flag_val and colors:
                bg, fg = colors
                flag_cell.fill = PatternFill("solid", fgColor=bg)
                flag_cell.font = Font(bold=True, color=fg)
                break

        wb.save(EXCEL_FILE_PATH)
    except Exception as e:
        logger.error(f"Failed to write news row: {e}")

# ─────────────────────────────────────────────
# EVENT FINGERPRINT LOAD
# ─────────────────────────────────────────────

def load_event_fingerprints(ws, logger):
    fingerprints = set()
    try:
        fp_col   = NEWS_HEADERS.index("Event Fingerprint") + 1
        date_col = NEWS_HEADERS.index("Date Found") + 1
        cutoff   = datetime.now() - timedelta(days=EVENT_DEDUP_DAYS)
        for row in ws.iter_rows(min_row=2, values_only=True):
            date_val = row[date_col - 1]
            fp_val   = row[fp_col - 1]
            if not date_val or not fp_val:
                continue
            try:
                row_date = (datetime.strptime(date_val, "%Y-%m-%d")
                            if isinstance(date_val, str) else date_val)
                if row_date >= cutoff:
                    fingerprints.add(fp_val)
            except Exception:
                continue
        logger.info(f"Event fingerprints loaded: {len(fingerprints)}")
    except Exception as e:
        logger.error(f"Error loading fingerprints: {e}")
    return fingerprints

# ─────────────────────────────────────────────
# OPENAI GPT PASS 1
# ─────────────────────────────────────────────

openai.api_key = OPENAI_API_KEY

def enrich_article(title, raw_summary, country, logger):
    prompt = f"""
You are a senior healthcare market intelligence analyst covering global emerging markets.
Your output feeds a B2B marketing team targeting hospitals, clinics, multispeciality clinics,
and daycare/early childhood centres across GCC, Africa, and Caribbean markets.

TARGET COUNTRY FOR THIS ARTICLE: {country}

DISCARD RULES — set relevance_score to "Low" if ANY apply:
- Article is primarily about a country OTHER than {country}
- Article is purely government health policy or budget with no specific named facility
- Facility is not a hospital, clinic, multispeciality clinic, or daycare/childcare/ECD centre
  (discard: schools, universities, veterinary, AI chatbot, virtual hospital, YMCA)
- Foreign institution covering {country} with no confirmed local presence or facility
- Article is about a financial loan or donor grant with no named physical facility
- Expansion is planned into {country} but no confirmed local presence yet
- Article is in French or another language AND covers a market not in target scope

MARKET CONTEXT FOR {country}:
- If {country} is in Africa: PPP awards, concession agreements, World Bank/AfDB/IFC
  funded projects, district hospitals, diagnostic centres, maternity centres, and ECD
  centres are all valid high-signal healthcare facility news
- If {country} is in the Caribbean: health centre groundbreakings, wellness centres,
  maternity units, and early childhood centre openings are valid signals
- If {country} is in GCC: standard hospital/clinic/daycare construction and investment news

MANAGEMENT NAME RULES:
- Extract the SHORT TRADING NAME the company is commonly known by
- Do NOT append legal suffixes (LLC, Ltd, WLL, Corp, Inc, SA, NV, GmbH)
- Do NOT append sub-brands or divisions (Academy, Foundation, Factory, Institute)
- Must be a proper noun with at least 2 words
- Also identify entity TYPE and extract website URL if mentioned

Headline: {title}
Summary: {raw_summary}

{NEWS_TYPE_GUIDE}
{RELEVANCE_GUIDE}

Respond ONLY with valid JSON — no markdown, no explanation.

{{
  "category": "Hospital" | "Clinic" | "Multispeciality Clinic" | "Daycare" | "Other",
  "news_type": "New Construction" | "Expansion" | "Acquisition / JV" | "Regulatory / Licensing" | "General News",
  "ai_summary": "2-3 sentence factual summary including country, facility type, and key figures.",
  "relevance_score": "High" | "Medium" | "Low",
  "relevance_reason": "One sentence citing which signals were found or absent.",
  "key_signal": "Single most actionable detail (investment figure, beds, timeline, award). null if none.",
  "management_name": "Short trading name — proper noun, 2+ words, no legal suffixes. null if unavailable.",
  "management_type": "Healthcare Operator" | "Developer / Contractor" | "Investment / Holding Group" | "Unknown",
  "company_website": "Website URL if mentioned. null if not mentioned."
}}
"""
    try:
        resp = openai.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.1, max_tokens=600,
        )
        raw = re.sub(r"```json|```", "", resp.choices[0].message.content).strip()
        return json.loads(raw)
    except Exception as e:
        logger.error(f"GPT Pass 1 error: {e}")
        return {
            "category":"Other","news_type":"General News",
            "ai_summary":raw_summary[:200],"relevance_score":"Medium",
            "relevance_reason":"Could not analyse.","key_signal":None,
            "management_name":None,"management_type":"Unknown",
            "company_website":None,
        }

# ─────────────────────────────────────────────
# GOOGLE NEWS
# ─────────────────────────────────────────────

def build_queries():
    """
    Build full query list per country using two-layer approach:
      Layer 1 — Universal queries (every country, every cluster)
      Layer 2 — Cluster-specific additive queries (based on MARKET_CLUSTERS)
    Deduplicates queries within the same country to avoid redundant calls.
    """
    queries = []
    for country in COUNTRIES:
        cluster   = MARKET_CLUSTERS.get(country, "gcc")
        seen_q    = set()   # dedup within this country

        # Layer 1 — Universal
        for category, templates in UNIVERSAL_QUERIES.items():
            for template in templates:
                q = template.format(country=country)
                if q not in seen_q:
                    seen_q.add(q)
                    queries.append((q, country, category))

        # Layer 2 — Cluster additive
        cluster_templates = CLUSTER_QUERIES.get(cluster, {})
        for category, templates in cluster_templates.items():
            for template in templates:
                q = template.format(country=country)
                if q not in seen_q:
                    seen_q.add(q)
                    queries.append((q, country, category))

    return queries

def fetch_google_news(query):
    encoded = requests.utils.quote(query)
    url = (f"https://news.google.com/rss/search"
           f"?q={encoded}&hl=en-US&gl=US&ceid=US:en")
    return feedparser.parse(url).entries

def make_hash(url, title):
    return hashlib.md5((url + title).encode("utf-8")).hexdigest()

def clean_html(text):
    return re.sub(r"<[^>]+>", "", text or "").strip()

def empty_person():
    return {"name":"","title":"","score":"","email":"","phone":"","linkedin":""}

# ─────────────────────────────────────────────
# MAIN PIPELINE
# ─────────────────────────────────────────────

def run():
    logger = setup_logging()
    logger.info("=" * 60)
    logger.info("GCC News Scraper started")
    logger.info("=" * 60)

    wb, ws_news, ws_apollo, ws_hashes = get_or_create_workbook()

    # Load all three dedup layers
    all_hashes           = load_processed_hashes(ws_news, ws_hashes, logger)
    event_fingerprints   = load_event_fingerprints(ws_news, logger)
    apollo_cache, apollo_row_index = load_apollo_cache(ws_apollo, logger)

    seen_hashes_this_run = set()
    seen_fps_this_run    = set()

    total_added = total_skipped = total_flagged = 0
    total_hash_skip = total_event_skip = 0
    total_cache_hits = total_apollo_calls = 0

    queries = build_queries()
    logger.info(f"Total queries this run: {len(queries)}")

    for query, country, category in queries:
        logger.info(f"Searching: {query}")
        try:
            entries = fetch_google_news(query)
        except Exception as e:
            logger.error(f"Google News fetch failed: {e}")
            continue

        for entry in entries:
            title       = entry.get("title","").strip()
            url         = entry.get("link","").strip()
            source      = entry.get("source",{}).get("title","Google News")
            raw_summary = clean_html(entry.get("summary",""))[:500]

            if not title or not url:
                continue

            # ── Layer 1: Hash check (pre-GPT) ──
            item_hash = make_hash(url, title)
            if item_hash in all_hashes or item_hash in seen_hashes_this_run:
                total_hash_skip += 1
                total_skipped   += 1
                continue

            # ── GPT Pass 1 ──
            logger.info(f"Analysing: {title[:70]}...")
            step1     = enrich_article(title, raw_summary, country, logger)
            relevance = step1.get("relevance_score","Medium")

            if relevance == "Low":
                logger.info(f"Skipped (Low) — {step1.get('relevance_reason','')}")
                write_processed_hash(wb, ws_hashes, item_hash, "Low", logger)
                seen_hashes_this_run.add(item_hash)
                all_hashes.add(item_hash)
                total_skipped += 1
                continue

            management_name = step1.get("management_name") or ""
            management_type = step1.get("management_type") or "Unknown"
            company_website = step1.get("company_website") or ""
            news_type       = step1.get("news_type","General News")

            # ── Layer 2: Event fingerprint ──
            fingerprint = make_event_fingerprint(
                management_name, country, news_type
            )
            if fingerprint in event_fingerprints or fingerprint in seen_fps_this_run:
                logger.info(
                    f"EVENT DEDUP — '{management_name}' | "
                    f"{country} | {news_type}"
                )
                write_processed_hash(wb, ws_hashes, item_hash,
                                     "Event Duplicate", logger)
                seen_hashes_this_run.add(item_hash)
                all_hashes.add(item_hash)
                total_event_skip += 1
                total_skipped    += 1
                continue

            research_flag  = "OK"
            company_fields = {}
            people         = []
            apollo_conf    = ""

            # ── Apollo enrichment (High only) ──
            if relevance == "High":
                if is_vague_name(management_name):
                    research_flag = "MANUAL RESEARCH NEEDED — management entity not identified"
                    total_flagged += 1
                else:
                    cache_key = apollo_cache_key(management_name, country)

                    if cache_key in apollo_cache:
                        # Layer 3 cache hit
                        cached         = apollo_cache[cache_key]
                        company_fields = cached["company"]
                        people         = cached["people"]
                        apollo_conf    = f"{cached['confidence']} (cached)"
                        total_cache_hits += 1
                        logger.info(f"Apollo CACHE HIT: '{management_name}'")
                    else:
                        # Fresh Apollo call
                        total_apollo_calls += 1
                        org, similarity, strategy = apollo_company_search(
                            management_name, company_website, country, logger
                        )
                        if org:
                            company_fields = extract_company_fields(org)
                            apollo_conf    = f"{int(similarity*100)}% via {strategy}"

                            # Gap 2 — Organisation Enrichment for size/description
                            domain = (extract_domain(company_fields.get("website"))
                                      or extract_domain(company_website))
                            if domain:
                                enriched_org = apollo_enrich_organisation(domain, logger)
                                company_fields["size"]        = enriched_org.get("size","")
                                company_fields["description"] = enriched_org.get("description","")

                            # People search + Gap 1/3/4 enrichment
                            people_raw = apollo_people_search(
                                org["id"], management_name, country, logger
                            )
                            top3_raw = prioritise_contacts(people_raw, logger)
                            # Full enrich each person
                            people = [extract_person_fields(p, logger)
                                      for p in top3_raw]

                            # Write to Apollo Results sheet
                            write_apollo_cache_row(
                                wb, ws_apollo, management_name, country,
                                company_fields, people, apollo_conf,
                                apollo_row_index, logger
                            )

                            # Update in-run cache
                            while len(people) < 3:
                                people.append(empty_person())
                            apollo_cache[cache_key] = {
                                "company":    company_fields,
                                "confidence": apollo_conf,
                                "people":     people,
                            }

                        elif similarity > 0:
                            research_flag = (
                                f"APOLLO MISMATCH — best score "
                                f"{int(similarity*100)}% below threshold"
                            )
                            apollo_conf = f"{int(similarity*100)}% (rejected)"
                            total_flagged += 1
                        else:
                            research_flag = "NOT IN APOLLO — manual research needed"
                            total_flagged += 1

                    # Gap 5 — Flag logic uses contact_has_data not just name
                    if research_flag == "OK":
                        if people and any(contact_has_data(p) for p in people):
                            if not any(p.get("name") for p in people):
                                research_flag = (
                                    "PARTIAL DATA — contact title/email found "
                                    "but name not available at API tier"
                                )
                        elif not people:
                            research_flag = (
                                "MANUAL RESEARCH NEEDED — "
                                "company found but no senior contacts"
                            )
                            total_flagged += 1

            # Pad to 3 contact slots
            while len(people) < 3:
                people.append(empty_person())

            row = [
                datetime.now().strftime("%Y-%m-%d"),
                title, source, url, country,
                step1.get("category","Other"),
                news_type,
                step1.get("ai_summary",""),
                relevance,
                step1.get("relevance_reason",""),
                step1.get("key_signal") or "",
                company_fields.get("name") or management_name,
                company_fields.get("website") or company_website,
                company_fields.get("linkedin",""),
                company_fields.get("size",""),
                company_fields.get("description",""),
                people[0]["name"], people[0]["title"], people[0].get("score",""),
                people[0]["email"], people[0]["phone"], people[0]["linkedin"],
                people[1]["name"], people[1]["title"], people[1].get("score",""),
                people[1]["email"], people[1]["phone"], people[1]["linkedin"],
                people[2]["name"], people[2]["title"], people[2].get("score",""),
                people[2]["email"], people[2]["phone"], people[2]["linkedin"],
                apollo_conf,
                research_flag,
                fingerprint,
                item_hash,
            ]

            write_news_row(wb, ws_news, row, logger)
            write_processed_hash(wb, ws_hashes, item_hash, "Written", logger)
            seen_hashes_this_run.add(item_hash)
            seen_fps_this_run.add(fingerprint)
            event_fingerprints.add(fingerprint)
            all_hashes.add(item_hash)
            total_added += 1
            logger.info(
                f"Saved | [{news_type}] | {relevance} | {management_type} | "
                f"Apollo: {apollo_conf or 'N/A'} | "
                f"Contacts: {sum(1 for p in people if contact_has_data(p))} | "
                f"Total: {total_added}"
            )

            time.sleep(ARTICLE_PAUSE)

    logger.info("=" * 60)
    logger.info(f"Run complete")
    logger.info(f"Added              : {total_added}")
    logger.info(f"Skipped (hash)     : {total_hash_skip}")
    logger.info(f"Skipped (event)    : {total_event_skip}")
    logger.info(f"Apollo cache hits  : {total_cache_hits}")
    logger.info(f"Apollo API calls   : {total_apollo_calls}")
    logger.info(f"Flagged            : {total_flagged}")
    logger.info(f"Excel : {EXCEL_FILE_PATH}")
    logger.info(f"Log   : {LOG_FILE_PATH}")
    logger.info("=" * 60)


if __name__ == "__main__":
    run()